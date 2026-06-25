# -*- coding: utf-8 -*-
"""파싱 결과 → 정리된 엑셀 리포트 생성.
시트: 안내 / 전체데이터 / 요약 / 불합격(NG) / 위험관리 / 추세
조건부 서식으로 공차 소진율을 색으로 표시(녹→황→적).
"""
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

COLS = [
    ("part_name", "부품명", 16), ("meas_date", "측정일", 11), ("block", "측정회차", 8),
    ("section", "공정/섹션", 22), ("dim_id", "측정ID", 9), ("char_type", "특성", 8),
    ("feature", "측정부위", 34), ("axis", "축", 5),
    ("nominal", "기준값", 11), ("tol_plus", "+공차", 8), ("tol_minus", "-공차", 8),
    ("meas", "측정값", 11), ("dev", "편차", 9), ("outtol", "공차초과", 9),
    ("tol_used_pct", "공차소진율%", 11), ("judge", "판정", 7), ("direction", "방향", 7),
]
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
NG_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def _style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}1"


def _write_df(ws, df, widths=None):
    ws.append(list(df.columns))
    for _, r in df.iterrows():
        ws.append(list(r))
    _style_header(ws, len(df.columns))
    for i, col in enumerate(df.columns, 1):
        w = (widths or {}).get(col, max(10, min(40, int(df[col].astype(str).str.len().max() if len(df) else 10) + 2)))
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_col=len(df.columns)):
        for cell in row:
            cell.border = THIN


def build_excel(rows, out_path):
    df = pd.DataFrame(rows)
    # 한글 컬럼명 매핑 (존재하는 것만)
    order = [c for c, _, _ in COLS if c in df.columns]
    df = df[order].copy()
    ko = {c: k for c, k, _ in COLS}
    widths_ko = {ko[c]: w for c, _, w in [(c, k, w) for c, k, w in COLS]}
    dfk = df.rename(columns=ko)

    from openpyxl import Workbook
    wb = Workbook()

    # 1) 안내 시트
    ws = wb.active; ws.title = "안내"
    guide = [
        ["3차원 측정데이터 AI 분석 리포트", ""],
        ["", ""],
        ["생성 방식", "CMM(pc-dmis) 측정 리포트 → 자동 파싱 → 정리/분석"],
        ["총 측정항목 수", len(df)],
        ["불합격(NG) 수", int((df["judge"] == "NG").sum())],
        ["위험(공차 70%+ 소진, OK) 수", int(((df["judge"] == "OK") & (df["tol_used_pct"] >= 70)).sum())],
        ["", ""],
        ["시트 안내", ""],
        ["전체데이터", "모든 측정항목(필터/정렬 가능)"],
        ["요약", "공정/특성별 합격·불합격 집계"],
        ["불합격(NG)", "규격을 벗어난 항목만"],
        ["위험관리", "합격이지만 공차의 70% 이상을 소진한 '예비 불량' 항목"],
        ["추세", "동일 부위가 여러 회차 측정된 항목(추세선 분석 대상)"],
        ["", ""],
        ["공차소진율%", "편차 ÷ 가까운 쪽 공차. 100% = 규격 한계, 100%↑ = 불합격"],
    ]
    for r in guide:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 60
    for r in (3, 4, 5, 6, 8):
        ws.cell(r, 1).font = Font(bold=True)

    # 2) 전체데이터
    _write_df(wb.create_sheet("전체데이터"), dfk, widths_ko)

    # 3) 요약 (공정별)
    g = (df.assign(NG=(df["judge"] == "NG").astype(int))
           .groupby("section")
           .agg(측정수=("judge", "size"), 불합격수=("NG", "sum"),
                평균소진율=("tol_used_pct", "mean"), 최대소진율=("tol_used_pct", "max"))
           .reset_index().rename(columns={"section": "공정/섹션"}))
    g["불합격률%"] = (g["불합격수"] / g["측정수"] * 100).round(1)
    g["평균소진율"] = g["평균소진율"].round(1); g["최대소진율"] = g["최대소진율"].round(1)
    g = g.sort_values("불합격수", ascending=False)
    _write_df(wb.create_sheet("요약"), g[["공정/섹션", "측정수", "불합격수", "불합격률%", "평균소진율", "최대소진율"]])

    # 4) 불합격
    ng = dfk[dfk["판정"] == "NG"].sort_values("공차초과", ascending=False)
    _write_df(wb.create_sheet("불합격(NG)"), ng, widths_ko)

    # 5) 위험관리
    risk = dfk[(dfk["판정"] == "OK") & (dfk["공차소진율%"] >= 70)].sort_values("공차소진율%", ascending=False)
    _write_df(wb.create_sheet("위험관리"), risk, widths_ko)

    # 6) 추세 (반복 측정 항목)
    key = df.groupby(["feature", "axis", "char_type"])
    rep_keys = key.size()[key.size() > 1].index
    trend = df.set_index(["feature", "axis", "char_type"]).loc[rep_keys].reset_index()
    trend = trend[["feature", "axis", "char_type", "block", "dim_id", "meas", "dev", "tol_used_pct", "judge"]]
    trend = trend.rename(columns=ko).sort_values(["측정부위", "축", "측정회차"])
    _write_df(wb.create_sheet("추세"), trend)

    # --- 조건부 서식 ---
    for sheet in ("전체데이터", "불합격(NG)", "위험관리"):
        ws = wb[sheet]
        cols = list(dfk.columns)
        # 공차소진율 색상 스케일
        if "공차소진율%" in cols:
            cl = get_column_letter(cols.index("공차소진율%") + 1)
            rng = f"{cl}2:{cl}{ws.max_row}"
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="num", start_value=0, start_color="63BE7B",
                mid_type="num", mid_value=70, mid_color="FFEB84",
                end_type="num", end_value=100, end_color="F8696B"))
        # 판정 NG 행 강조
        if "판정" in cols:
            jl = get_column_letter(cols.index("판정") + 1)
            ws.conditional_formatting.add(
                f"{jl}2:{jl}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"NG"'], fill=NG_FILL))

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import sys
    from parse_cmm import parse_file
    rows = parse_file(sys.argv[1] if len(sys.argv) > 1 else "3d")
    out = build_excel(rows, "out/측정데이터_정리.xlsx")
    print("저장:", out)
